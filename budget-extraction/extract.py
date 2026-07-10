#!/usr/bin/env python3
"""
Budget.xlsx -> Budgeteer extractor.

Implements EXTRACTION_ROADMAP.md (Rev 2 + Rule 4). Emits:
  - budgeteer_import.json   (BudgeteerBackup-shaped: tables{} of raw rows, integer cents)
  - EXTRACTION_REPORT.md    (validation + discrepancy register, computed live)

Pipeline (see roadmap §5):
  1. Parse every tab; split Deposit/Withdrawal formulas into terms paired 1:1 with description items.
  2. Classify: Reset (Rule 4) -> Envelope transfer (Rule 3) -> deposit pool / withdrawal pool.
  3. Consolidate same-week/same-source deposits into one transaction w/ multi-envelope allocations (Rule 2).
  4. One transaction+allocation per remaining withdrawal term (Rule 1).
  5. Reset -> one account withdrawal + per-envelope allocations (Rule 4).
"""
import openpyxl, re, json, uuid, sys
from collections import defaultdict, Counter
from datetime import datetime, timezone

SRC = "Budget.xlsx"
DEFAULT_HOUSEHOLD_ID = "00000000-0000-0000-0000-000000000001"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")   # deterministic id namespace
EPOCH = "2014-07-18T00:00:00Z"
ACCOUNT_NAME = "Budget"
ACCOUNT_KIND = "checking"
ARCHIVE_PREFIX = re.compile(r'^\s*archive\s*', re.IGNORECASE)
TRANSFER_WORDS = ("transfer", "left in saving", "from saving", "to saving")

def uid(*parts):
    return str(uuid.uuid5(NS, "|".join(str(p) for p in parts)))

def cents(x):
    return int(round(float(x) * 100))

def iso_ts(d):
    return d.strftime("%Y-%m-%dT00:00:00Z")

def iso_date(d):
    return d.strftime("%Y-%m-%d")

def split_terms(f):
    """Split a sum formula into top-level terms; a (..) group stays one term."""
    if not isinstance(f, str):
        return [str(f)]
    s = f.strip().lstrip("=")
    out, depth, cur = [], 0, ""
    for ch in s:
        if ch == "(":
            depth += 1; cur += ch
        elif ch == ")":
            depth -= 1; cur += ch
        elif ch == "+" and depth == 0:
            if cur.strip(): out.append(cur.strip()); cur = ""
        else:
            cur += ch
    if cur.strip(): out.append(cur.strip())
    return out

def ev(t):
    try:
        return float(eval(str(t), {"__builtins__": {}}, {}))
    except Exception:
        return None

def desc_items(d):
    return [x.strip() for x in str(d).split(",")] if d not in (None, "") else []

def norm(s):
    return re.sub(r"\s+", " ", str(s).strip()).lower()

def is_transfer(label):
    return any(w in str(label).lower() for w in TRANSFER_WORDS)

# ---------------------------------------------------------------- parse
wb = openpyxl.load_workbook(SRC, data_only=False)
wbv = openpyxl.load_workbook(SRC, data_only=True)

report = {"warnings": [], "mismatch_wd": 0, "mismatch_dep": 0, "bad_eval": [],
          "neg_terms": 0, "zero_terms": 0, "merged_dep_allocs": 0}

# term record: dict(sheet, week(date), label, cents, is_tf)
dep_terms, wd_terms, reset_allocs = [], [], []

for name in wb.sheetnames:
    ws = wb[name]
    for r in range(2, ws.max_row + 1):
        dcell = ws.cell(row=r, column=1).value
        wk = dcell.date() if hasattr(dcell, "year") else None
        depD = ws.cell(row=r, column=5).value
        wv = ws.cell(row=r, column=3).value

        # Rule 4: reset row (label "Reset" in Deposit Description col, value is a withdrawal)
        if isinstance(depD, str) and depD.strip().lower() == "reset" and wv not in (None, "", 0):
            terms = split_terms(wv)
            amt = sum(ev(t) for t in terms if ev(t) is not None)
            reset_allocs.append({"sheet": name, "week": wk, "cents": cents(amt)})
            continue

        for col, bucket in ((2, dep_terms), (3, wd_terms)):
            v = ws.cell(row=r, column=col).value
            if v in (None, "", 0):
                continue
            items = desc_items(ws.cell(row=r, column=col + 3).value)
            terms = split_terms(v)
            if len(terms) != len(items):
                report["mismatch_wd" if col == 3 else "mismatch_dep"] += 1
            overflow = ", ".join(items[len(terms):]) if len(items) > len(terms) else None
            for i, t in enumerate(terms):
                a = ev(t)
                if a is None:
                    report["bad_eval"].append((name, r, col, t)); continue
                if a < 0: report["neg_terms"] += 1
                if a == 0: report["zero_terms"] += 1
                if i < len(items):
                    label = items[i]
                elif col == 2 and len(items) == 1:
                    label = items[0]                 # single source funds a multi-term deposit
                else:
                    label = ""                       # extra withdrawal term -> blank payee
                rec = {"sheet": name, "week": wk, "label": label, "cents": cents(a),
                       "is_tf": is_transfer(label)}
                # attach description overflow to the final term's memo
                if overflow and i == len(terms) - 1:
                    rec["overflow"] = overflow
                bucket.append(rec)

# ---------------------------------------------------------------- Rule 3: envelope transfers
dep_tf_by_week = defaultdict(list)
for idx, d in enumerate(dep_terms):
    if d["is_tf"] and d["week"] is not None:
        dep_tf_by_week[d["week"]].append(idx)

used_dep = set()
env_transfers = []          # (week, from_sheet, to_sheet, cents)
wd_consumed = set()
for wi, w in enumerate(wd_terms):
    if not (w["is_tf"] and w["week"] is not None):
        continue
    for di in dep_tf_by_week.get(w["week"], []):
        if di in used_dep:
            continue
        d = dep_terms[di]
        if d["sheet"] != w["sheet"] and d["cents"] == w["cents"]:
            used_dep.add(di); wd_consumed.add(wi)
            env_transfers.append((w["week"], w["sheet"], d["sheet"], w["cents"]))
            break

# ---------------------------------------------------------------- envelopes (27)
def env_name(sheet):
    return ARCHIVE_PREFIX.sub("", sheet).strip() if ARCHIVE_PREFIX.match(sheet) else sheet.strip()

def last_activity_date(sheet):
    ws = wbv[sheet]; last = None
    for r in range(2, ws.max_row + 1):
        dv = ws.cell(row=r, column=2).value; wv = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=1).value
        num = lambda v: isinstance(v, (int, float)) and not isinstance(v, bool) and v != 0
        if (num(dv) or num(wv)) and hasattr(d, "year"):
            last = d
    return last

def first_activity_date(sheet):
    ws = wbv[sheet]
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if hasattr(d, "year"):
            return d
    return None

sheet_to_env_id, envelopes_rows = {}, []
for sheet in wb.sheetnames:
    nm = env_name(sheet)
    eid = uid("envelope", norm(nm))
    sheet_to_env_id[sheet] = eid
    archived = ARCHIVE_PREFIX.match(sheet) is not None
    fa = first_activity_date(sheet)
    envelopes_rows.append({
        "id": eid, "household_id": DEFAULT_HOUSEHOLD_ID, "name": nm, "kind": "standard",
        "created_at": iso_ts(fa) if fa else EPOCH,
        "archived_at": iso_ts(last_activity_date(sheet)) if archived else None,
    })

# ---------------------------------------------------------------- account (1)
account_id = uid("account", norm(ACCOUNT_NAME))
accounts_rows = [{
    "id": account_id, "household_id": DEFAULT_HOUSEHOLD_ID, "name": ACCOUNT_NAME,
    "kind": ACCOUNT_KIND, "created_at": EPOCH, "archived_at": None,
}]

# ---------------------------------------------------------------- transactions + allocations
txns, allocs = [], []
seq = 0
def new_txn(amount_cents, occurred, payee, memo):
    global seq
    seq += 1
    tid = uid("txn", seq)
    txns.append({
        "id": tid, "household_id": DEFAULT_HOUSEHOLD_ID, "account_id": account_id,
        "amount_cents": amount_cents, "kind": "normal", "occurred_on": occurred,
        "payee": payee or None, "memo": memo or None,
        "transfer_id": None, "recurring_id": None, "created_at": iso_ts_from_iso(occurred),
    })
    return tid

def iso_ts_from_iso(d):  # d is 'YYYY-MM-DD'
    return d + "T00:00:00Z"

def add_alloc(tid, env_id, amount_cents):
    allocs.append({"id": uid("alloc", tid, env_id, len(allocs)),
                   "transaction_id": tid, "envelope_id": env_id, "amount_cents": amount_cents})

# Deposits: group remaining by (week, normalized source); merge same-envelope dupes
dep_groups = defaultdict(lambda: defaultdict(int))   # (week,src) -> {sheet: cents}
dep_group_payee = {}                                 # (week,src) -> representative original label
dep_group_memo = {}
for idx, d in enumerate(dep_terms):
    if idx in used_dep:
        continue
    src = norm(d["label"]) if d["label"].strip() else "(unlabeled)"
    key = (d["week"], src)
    before = len(dep_groups[key])
    dep_groups[key][d["sheet"]] += d["cents"]
    if len(dep_groups[key]) == before and d["sheet"] in dep_groups[key]:
        report["merged_dep_allocs"] += 0  # placeholder; counted below
    dep_group_payee.setdefault(key, d["label"].strip() or None)
    if d.get("overflow"):
        dep_group_memo[key] = d["overflow"]

dep_alloc_terms_before_merge = sum(1 for i, _ in enumerate(dep_terms) if i not in used_dep)
dep_alloc_rows = sum(len(v) for v in dep_groups.values())
report["merged_dep_allocs"] = dep_alloc_terms_before_merge - dep_alloc_rows

for (week, src), env_amounts in dep_groups.items():
    occurred = iso_date(week)
    total = sum(env_amounts.values())
    payee = dep_group_payee.get((week, src))
    tid = new_txn(total, occurred, payee, dep_group_memo.get((week, src)))
    for sheet, c in env_amounts.items():
        add_alloc(tid, sheet_to_env_id[sheet], c)

# Withdrawals: one txn + one alloc per remaining term
for wi, w in enumerate(wd_terms):
    if wi in wd_consumed:
        continue
    occurred = iso_date(w["week"])
    tid = new_txn(-w["cents"], occurred, w["label"] or None, w.get("overflow"))
    add_alloc(tid, sheet_to_env_id[w["sheet"]], -w["cents"])

# Reset: ONE withdrawal txn + per-envelope allocations
reset_total = sum(a["cents"] for a in reset_allocs)
reset_week = reset_allocs[0]["week"] if reset_allocs else None
if reset_allocs:
    occurred = iso_date(reset_week)
    tid = new_txn(-reset_total, occurred, "Reset", "End-of-budget reset (zero-out)")
    for a in reset_allocs:
        add_alloc(tid, sheet_to_env_id[a["sheet"]], -a["cents"])

# ---------------------------------------------------------------- envelope_transfers rows
env_transfer_rows = []
for i, (week, frm, to, c) in enumerate(env_transfers):
    env_transfer_rows.append({
        "id": uid("envxfer", i), "household_id": DEFAULT_HOUSEHOLD_ID,
        "from_envelope_id": sheet_to_env_id[frm], "to_envelope_id": sheet_to_env_id[to],
        "amount_cents": c, "occurred_on": iso_date(week), "memo": "Transfer",
        "created_at": iso_ts(week),
    })

# ---------------------------------------------------------------- assemble BudgeteerBackup
backup = {
    "meta": {
        "schema_name": "budgeteer", "schema_version": "V1",
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": SRC, "money_unit": "integer minor units (US cents)",
        "household_id": DEFAULT_HOUSEHOLD_ID,
    },
    "tables": {
        "households": [{"id": DEFAULT_HOUSEHOLD_ID, "name": "Default household", "created_at": EPOCH}],
        "accounts": accounts_rows,
        "envelopes": envelopes_rows,
        "transfers": [],
        "envelope_transfers": env_transfer_rows,
        "templates": [], "template_lines": [],
        "recurring_transactions": [], "recurring_lines": [],
        "transactions": txns,
        "allocations": allocs,
        "reconciliations": [],
        "envelope_targets": [], "credit_limits": [], "loan_principals": [],
    },
}

with open("budgeteer_import.json", "w") as f:
    json.dump(backup, f, indent=2)

# ---------------------------------------------------------------- validation / cross-checks
def num(v): return isinstance(v, (int, float)) and not isinstance(v, bool)

# per-envelope derived balance from our generated rows:
#   Σ allocations  +  net envelope-transfer flow (to gains, from loses)
env_balance = defaultdict(int)
for a in allocs:
    env_balance[a["envelope_id"]] += a["amount_cents"]
for t in env_transfer_rows:
    env_balance[t["from_envelope_id"]] -= t["amount_cents"]
    env_balance[t["to_envelope_id"]] += t["amount_cents"]

# sheet final balance (data_only)
sheet_final = {}
for sheet in wb.sheetnames:
    ws = wbv[sheet]; last = 0
    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=4).value
        if num(b): last = b
    sheet_final[sheet] = cents(last)

xcheck_fail = []
for sheet in wb.sheetnames:
    eid = sheet_to_env_id[sheet]
    got = env_balance.get(eid, 0)
    want = sheet_final[sheet]
    if abs(got - want) > 1:    # 1 cent tolerance
        xcheck_fail.append((sheet, got, want))

# account balance = sum of all txns
account_balance = sum(t["amount_cents"] for t in txns)

# split invariant check
alloc_by_txn = defaultdict(int)
for a in allocs:
    alloc_by_txn[a["transaction_id"]] += a["amount_cents"]
inv_fail = []
for t in txns:
    s = alloc_by_txn[t["id"]]
    if abs(s) > abs(t["amount_cents"]) + 0 or (s != 0 and (s > 0) != (t["amount_cents"] > 0)):
        inv_fail.append((t["id"], t["amount_cents"], s))

n_dep_txn = len(dep_groups)
n_wd_txn = sum(1 for wi in range(len(wd_terms)) if wi not in wd_consumed)
n_reset_txn = 1 if reset_allocs else 0

summary = f"""# Extraction Report — Budget.xlsx → Budgeteer

Generated: {backup['meta']['generated_at']}  ·  Source: `{SRC}`

## Generated entity counts
| Entity | Rows |
|--------|-----:|
| households | {len(backup['tables']['households'])} |
| accounts | {len(accounts_rows)} |
| envelopes | {len(envelopes_rows)}  ({sum(1 for e in envelopes_rows if e['archived_at'])} archived) |
| transactions | {len(txns)}  ({n_dep_txn} deposit + {n_wd_txn} withdrawal + {n_reset_txn} reset) |
| allocations | {len(allocs)} |
| envelope_transfers | {len(env_transfer_rows)} |
| (all other tables) | 0 |

## Money reconciliation (integer cents)
- Account balance (Σ all transactions): **{account_balance} cents** (expected 0)
- Reset transaction total: **-{reset_total} cents** (${reset_total/100:,.2f}) across {len(reset_allocs)} envelopes
- Envelope-transfer total: **{sum(t['amount_cents'] for t in env_transfer_rows)} cents** (${sum(t['amount_cents'] for t in env_transfer_rows)/100:,.2f}) over {len(env_transfer_rows)} transfers

## Discrepancy register
- Term↔description count mismatches — withdrawals: **{report['mismatch_wd']}**, deposits: **{report['mismatch_dep']}** (positional fallback applied; amounts preserved)
- Non-evaluable terms (cross-sheet refs, skipped): **{len(report['bad_eval'])}** → {report['bad_eval']}
- Negative-valued terms: **{report['neg_terms']}**  ·  Zero-valued terms: **{report['zero_terms']}**
- Merged deposit allocations (same week+source+envelope): **{report['merged_dep_allocs']}**

## Cross-checks (must pass)
- Split invariant violations: **{len(inv_fail)}** {inv_fail[:5]}
- Per-envelope balance mismatches (generated vs sheet final Balance): **{len(xcheck_fail)}**
{chr(10).join(f'    - {s}: got {g} want {w} (Δ {g-w})' for s,g,w in xcheck_fail) if xcheck_fail else '    (all 27 envelopes reconcile within 1¢)'}

## Archived envelopes
{chr(10).join(f"    - {e['name']}: archived_at {e['archived_at'][:10]}" for e in envelopes_rows if e['archived_at'])}
"""
with open("EXTRACTION_REPORT.md", "w") as f:
    f.write(summary)

print(summary)
print(f"\nWrote budgeteer_import.json ({len(json.dumps(backup))/1e6:.2f} MB) and EXTRACTION_REPORT.md")
